# -*- coding: utf-8 -*-
"""生成三套数据集的字段矩阵（建表依据）。"""
import openpyxl, os, json, re

SP = "/tmp/claude-1000/-home-joii-project-cpq/0ba29c9c-75e6-47a4-ab46-934bb746ffa9/scratchpad"
D  = "dev-docs/task-260902-报价与核价建表与导入方案新规范/"

DATASETS = [
    ("QUOTE",       "报价数据",   "ds_quote_",       "报价 - 数据导入与表格建表.xlsx"),
    ("COST_BASIC",  "基础核价",   "ds_cost_basic_",  "核价2 - 数据导入与表格建表.xlsx"),
    ("COST_DETAIL", "详细核价",   "ds_cost_detail_", "核价1 - 数据导入与表格建表.xlsx"),
]

# ── 中文列名 → 英文字段名（沿用现有 V6 命名习惯）────────────────────────────
COL = {
 "销售料号":"material_no","生产料号":"production_no","品名":"material_name","规格":"specification",
 "尺寸":"dimension","旧料号":"old_material_no","单重":"unit_weight","项次":"item_seq",
 "客户编号":"customer_no","客户料号名称":"customer_part_name","客户产品编号":"customer_product_no","客户图号":"customer_drawing_no",
 "投入类型":"input_type","投入料号":"input_material_no","投入料号名称":"input_material_name",
 "产出料号类型":"output_material_type","组成数量":"component_qty","材料毛重":"gross_weight",
 "材料净重":"net_weight","重量单位":"weight_unit","材料占比（%）":"material_ratio",
 "损耗率（%）":"loss_rate","损耗率%":"loss_rate","不良率（%）":"defect_rate",
 "材质料号":"material_part_no","材质料号名称":"material_part_name","元素":"element_code",
 "元素代码":"element_code","组成含量（%）":"content_pct","毛用量":"gross_usage",
 "毛用量单位":"gross_usage_unit","净用量":"net_usage","净用量单位":"net_usage_unit",
 "回收折扣(%)":"recovery_discount","回收折扣（%）":"recovery_discount","回收量":"recovery_qty",
 "基准值":"base_value","比例（%）":"ratio_pct","货币":"currency","计价单位":"pricing_unit",
 "是否随材料价格波动":"follow_material_price","材料结算涨幅比例（%）":"material_increase_ratio",
 "材料固定的涨幅值":"material_increase_value","涨幅货币":"increase_currency","涨幅单位":"increase_unit",
 "要素项次":"element_item_seq","要素名称":"element_name","值":"value",
 "回收值":"recovery_value","回收来源":"recovery_source",
 "工序项次":"operation_item_seq","工序编号":"operation_no","工序名称":"operation_name",
 "组成件料号":"sub_component_no","组成件名称":"sub_component_name",
 "供应商编号":"supplier_no","供应商名称":"supplier_name",
 "组装工序":"assembly_operation","组装加工费":"assembly_fee","拒收率/不良率（%）":"defect_rate",
 "年降顺序":"discount_seq","年降系数（%）":"discount_rate","年降系数（%/年）":"discount_rate",
 "单次固定年降值":"fixed_discount_value","单次固定年降金额":"fixed_discount_value","降价次数":"discount_times",
 "电镀方案编号":"plating_scheme_no","版本编号":"plating_version","电镀加工费":"plating_process_fee",
 "电镀材料费":"plating_material_fee","方案编号":"scheme_no","版本":"scheme_version",
 "电镀元素名称":"plating_element","元素单价来源网站网址":"price_source_url",
 "元素单价来源网站名称":"price_source_name","元素单价抓取规则":"price_fetch_rule",
 "电镀面积（cm2）":"plating_area","镀层厚度（μm）":"coating_thickness","电镀要求":"plating_requirement",
 "密度（g/cm3)":"density",
 "组成类型":"component_type","组成料号":"component_no","使用特性":"usage_characteristic",
 "组成用量":"component_qty","组成用量单位":"component_qty_unit","底数":"base_qty","底数单位":"base_qty_unit",
 "材料损耗率（%）":"material_loss_rate","材料固定损耗量":"material_fixed_loss",
 "人工标准单价":"labor_std_price","币种":"currency","计量单位":"unit",
 "折旧单价":"depreciation_price","生产能耗单价":"production_energy_price",
 "非生产能耗单价":"auxiliary_energy_price","模具台账/工装编号":"tooling_no",
 "单个模具/工装成本":"tooling_cost","寿命（次）":"tooling_life","单循环产量":"cycle_output",
 "模具工装成本单价":"tooling_unit_price","耗材成本单价":"consumable_price",
 "包装成本单价":"packaging_price","来料料号":"incoming_material_no","加工费":"process_fee",
 "损耗（%）":"loss_rate","费用":"fee","外加工费用":"outsourced_fee","单位":"unit",
 "不良率/拒收率（%）":"defect_rate",
}

# ── sheet 名 → 表名后缀 ────────────────────────────────────────────────────
TBL = {
 "物料":"material","客户料号":"customer_part","物料BOM":"material_bom",
 "物料与元素BOM":"element_bom","来料固定加工费":"incoming_fixed_fee","来料其他费用":"incoming_other_fee",
 "来料回收折扣":"incoming_recovery","自制加工费":"self_process_fee","成品其他费用":"finished_other_fee",
 "组成件其他费用":"sub_component_fee","组装加工费":"assembly_fee","组装加工费年降":"assembly_fee_annual",
 "电镀费用":"plating_fee","电镀方案":"plating_scheme","来料年降":"incoming_annual","年降系数":"annual_discount",
 "产能":"capacity","设备折旧成本":"depreciation","生产设备能耗":"production_energy",
 "辅助设备能耗":"auxiliary_energy","模具工装成本":"tooling","生产耗材BOM":"consumable",
 "包装材料BOM":"packaging","来料加工费":"incoming_process_fee","来料其他固定费用":"incoming_other_fixed_fee",
 "加工费&组装费":"process_assembly_fee","其他外加工成本":"outsourced_process","电镀成本":"plating_cost",
 "成品其他比例费用":"finished_ratio_fee","成品其他固定费用":"finished_fixed_fee",
}

# ── PG 类型推断（2026-09-03 重写）────────────────────────────────────────────
# 🚩 原版按关键词单层匹配，把「重量单位」「计价单位」「毛用量单位」「元素单价来源网站网址」
#    这类列误判成 numeric（「重」「价」「用量」命中在先），实测 25 列错误，照建会让导入必然
#    报「不是合法数值」。改为**分层判定，文本型规则优先于数值型规则**。
INT_KW  = ("项次", "顺序", "次数", "寿命", "产量")
DEC_KW  = ("率", "比例", "占比", "含量", "重", "价", "费", "成本", "金额", "值",
           "用量", "数量", "底数", "面积", "厚度", "密度", "折扣", "损耗")

def pgtype(cn):
    """PG 类型推断（2026-09-03 修订 v2）。

    🚩 v1 按关键词单层匹配，把「重量单位」「计价单位」「毛用量单位」「元素单价来源网站网址」
       误判成 numeric（「重」「价」「用量」命中在先），且「损耗（%）」「材料固定损耗量」
       「年降系数（%）」漏判成 varchar —— 实测由后端 #1 拿 Excel 原始 cell 值反查发现。
    🚩 v2 原则：**只修类型方向，不做无证据的长度收紧**。
       varchar 默认仍是 128（与 v1 一致），只有三类有明确依据的例外：
         · 客户编号 20  —— 与现有 material_customer_map.customer_no varchar(20) 对齐
         · 网址     512 —— URL 可能长
         · 名称/要求/规则/来源 256 —— 中文长名
    """
    c = cn.strip()
    if c == "客户编号":                            return "varchar(20)"    # D-18，对齐现有 mcm
    if c.startswith("是否"):                       return "boolean"
    # ── 文本型硬规则：必须先判，否则被下面的数值关键词抢走 ──
    if c.endswith("单位"):                         return "varchar(128)"   # 重量/计价/毛用量/净用量/组成用量/底数/涨幅单位
    if "网址" in c:                                return "varchar(512)"
    if any(k in c for k in ("名称", "规则", "来源", "要求")):  return "varchar(256)"
    if any(k in c for k in ("编号", "料号", "图号", "代码", "符号",
                            "类型", "特性", "币种", "货币")):   return "varchar(128)"
    # ── 百分比一律数值（v1 漏判「损耗（%）」「年降系数（%）」）──
    if "%" in c or "％" in c:                      return "numeric(26,12)"
    if any(k in c for k in INT_KW):                return "integer"
    if any(k in c for k in DEC_KW):                return "numeric(26,12)"
    return "varchar(128)"

def sem(c):
    if c == 'FFFF0000': return 'REQ'
    if c in ('theme0','none'): return 'SKIP'
    return 'OPT'

def color_of(cell):
    try:
        fg = cell.fill.fgColor
        if fg.type=='rgb' and fg.rgb not in (None,'00000000','FFFFFFFF'): return fg.rgb
        if fg.type=='theme': return "theme%s"%fg.theme
    except Exception: pass
    return "none"

def load(fn):
    wb = openpyxl.load_workbook(os.path.join(D, fn))
    out=[]
    for ws in wb.worksheets:
        r2=[ws.cell(2,c).value for c in range(1,ws.max_column+1)]
        versioned=any(v in ('轴','对比项') for v in r2 if v)
        cols=[]
        for c in range(1, ws.max_column+1):
            n=ws.cell(1,c).value
            if n is None: continue
            m=ws.cell(2,c).value if versioned else None
            cols.append([str(n).strip(), color_of(ws.cell(1,c)), m if m in ('轴','对比项') else None])
        out.append({"name":ws.title,"versioned":versioned,"cols":cols})
    return out

data={}
for key,label,prefix,fn in DATASETS:
    try:
        data[key]=load(fn); src="机器解析"
    except Exception as e:
        data[key]=json.load(open(os.path.join(SP,"detail_fallback.json"),encoding="utf-8")); src="人工转录（文件损坏）"
    data[key+"__src"]=src

json.dump(data, open(os.path.join(SP,"matrix_full.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)

# ── 生成 md ────────────────────────────────────────────────────────────────
L=[]
L.append("# 附录 · 三套数据集字段矩阵（建表唯一依据）\n")
L.append("> 由 `scratchpad/gen_matrix.py` 从三份 Excel 机器生成。**改 Excel 必须重跑本脚本并更新本文件**。\n")
L.append("> 底色语义：🔴 红 `FFFF0000` = 必填建字段 · ⚪ 白/无填充 `theme0` = **不建字段**（主数据 JOIN 展示列）· 🟡 其余色 = 选填建字段。\n")
L.append("> 轴列**无论底色一律必填**（闸门 A0 裁决）。\n")
missing=set()
# ── D-18：报价「客户料号」补客户编号列（Excel 尚未补，此处固化）──
for _s in data.get("QUOTE", []):
    if _s["name"] == "客户料号" and not any(c[0] == "客户编号" for c in _s["cols"]):
        _s["cols"].insert(0, ["客户编号", "FFFF0000", None])

for key,label,prefix,fn in DATASETS:
    sheets=data[key]
    nv=sum(1 for s in sheets if s["versioned"])
    L.append("\n---\n\n## %s（`%s*`）· %s\n"%(label,prefix,data[key+'__src']))
    L.append("来源：`%s` · sheet %d 张（带版本 %d / 免版本 %d）\n"%(fn,len(sheets),nv,len(sheets)-nv))
    for s in sheets:
        suffix=TBL.get(s["name"],"TODO_"+re.sub(r'\W','_',s["name"]))
        tbl=prefix+suffix
        axis=[c[0] for c in s["cols"] if c[2]=="轴"]
        L.append("\n### %s → `%s`%s\n"%(s["name"],tbl," 🕐 **带版本** + `%s_history`"%tbl if s["versioned"] else " 🔓 免版本"))
        if s["versioned"]:
            L.append("轴：**%s** ｜ 比对项 %d 列\n"%(axis[0] if axis else "⚠️缺",sum(1 for c in s['cols'] if c[2]=='对比项')))
        L.append("\n| # | Excel 列名 | 底色 | 建字段 | 字段名 | PG 类型 | 标记 |")
        L.append("|---|---|---|---|---|---|---|")
        i=0
        for cn,fill,mark in s["cols"]:
            sm=sem(fill)
            if sm=="SKIP":
                L.append("| — | %s | ⚪ %s | ❌ 不建 | — | — | 主数据带出 |"%(cn,fill)); continue
            i+=1
            en=COL.get(cn)
            if en is None: missing.add(cn); en="⚠️TODO"
            icon="🔴 必填" if sm=="REQ" else "🟡 选填"
            if mark=="轴": icon="🔴 必填(轴)"
            L.append("| %d | %s | %s | ✅ | `%s` | %s | %s |"%(i,cn,fill,en,pgtype(cn),mark or "—"))
        L.append("")
if missing:
    L.append("\n---\n\n## ⚠️ 未命中英文名映射的列（需后端补）\n")
    for m in sorted(missing): L.append("- `%s`"%m)

open(os.path.join(D,"字段矩阵.md"),"w",encoding="utf-8").write("\n".join(L))
print("已生成 字段矩阵.md，行数 %d，未映射列 %d 个：%s"%(len(L),len(missing),"、".join(sorted(missing)) or "无"))
